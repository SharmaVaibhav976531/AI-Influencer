import csv
import io
import logging
from datetime import datetime
from django.http import StreamingHttpResponse, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from apps.influencers.services.result_service import get_filtered_classifications
from apps.classification.models import Classification

logger = logging.getLogger(__name__)

EXPORT_HEADERS = [
    "Name", "Handle", "Platform", "Followers", "Following", "Posts",
    "Language", "Location", "Bio", "Description",
    "Overall Score", "Confidence Score", "Recommendation", "Orientation",
    "Rule-Based Score", "Matched Keywords", "Government Scheme Mentions",
    "Development Topics", "Summary", "Reason",
    "Upload File", "Upload Date", "Classification Date", "Processing Status"
]

def get_export_queryset(request, export_type):
    """Builds the queryset for export based on the requested type."""
    if export_type == 'selected':
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            return Classification.objects.none()
        return Classification.objects.filter(
            id__in=selected_ids,
            influencer__upload__user=request.user,
            status='COMPLETED'
        ).select_related('influencer', 'influencer__upload')
    
    # For 'all', 'filtered', or 'page', reuse the existing filter logic
    # Note: 'page' is handled by the view passing only the current page's IDs or 
    # we can just export the filtered set. For simplicity, 'page' exports the filtered set 
    # unless we specifically pass page IDs. We'll treat 'page' as 'filtered' for backend 
    # simplicity, or the view can pass a 'page_ids' list. Let's support 'filtered' and 'selected'.
    return get_filtered_classifications(request.user, request.GET)

def format_row(classification):
    """Formats a single Classification object into a list of export-ready values."""
    inf = classification.influencer
    ai_resp = classification.ai_response or {}
    
    return [
        inf.name,
        inf.handle,
        inf.get_platform_display(),
        inf.followers,
        inf.following,
        inf.total_posts,
        inf.language_detected or "Unknown",
        inf.location or "",
        inf.bio or "",
        inf.description or "",
        float(classification.overall_score) if classification.overall_score else 0.0,
        float(classification.confidence_score) if classification.confidence_score else 0.0,
        classification.get_recommendation_display(),
        "Supportive" if classification.orientation_match else "Neutral/Unknown",
        float(inf.rule_based_score) if inf.rule_based_score else 0.0,
        ", ".join(classification.matched_keywords) if classification.matched_keywords else "",
        ", ".join(ai_resp.get("government_scheme_mentions", [])),
        ", ".join(ai_resp.get("development_topics", [])),
        classification.summary or "",
        classification.reason or "",
        inf.upload.original_filename,
        inf.upload.created_at.strftime("%Y-%m-%d %H:%M") if inf.upload.created_at else "",
        classification.created_at.strftime("%Y-%m-%d %H:%M") if classification.created_at else "",
        classification.get_status_display()
    ]

def generate_csv_response(queryset, filename):
    """Generates a streaming CSV response with UTF-8 BOM for Excel compatibility."""
    def iterator():
        # UTF-8 BOM ensures Excel opens Hindi/Unicode characters correctly
        yield '\ufeff'
        
        writer = csv.writer(io.StringIO())
        writer.writerow(EXPORT_HEADERS)
        yield writer.writerow
        
        for obj in queryset.iterator(chunk_size=1000):
            row = format_row(obj)
            # We need to yield the actual string, so we use a dummy StringIO per row 
            # or just yield the joined string. Better: yield from a fresh StringIO.
            output = io.StringIO()
            w = csv.writer(output)
            w.writerow(row)
            yield output.getvalue()

    response = StreamingHttpResponse(iterator(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def generate_excel_response(queryset, filename):
    """Generates an Excel response with professional formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Influencer Results"
    
    # Write Headers
    ws.append(EXPORT_HEADERS)
    
    # Style Headers
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write Data
    for obj in queryset.iterator(chunk_size=1000):
        ws.append(format_row(obj))
    
    # Formatting
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    
    # Auto-size columns (with a max limit to prevent absurd widths)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50) # Max 50 characters width
        ws.column_dimensions[column_letter].width = adjusted_width
        
        # Enable text wrapping for all cells in this column
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response